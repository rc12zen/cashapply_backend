"""
app.bank_statement.parser
==========================
Generic credit row parser driven by bank_configs.json.

Pipeline per file
-----------------
1. ExtractorFactory.extract()  — format-specific DataFrame
2. validate_columns()          — LOUD error if any required column is missing
3. resolve_file_fields()       — cell / fixed / filename_pattern (once per file)
4. apply_exclusions()          — drop rows that match exclusion rules
5. Row loop:
   a. eval_credit_rule()       — skip non-credit rows
   b. resolve_row_fields()     — column / concat per row
   c. apply_transforms()       — clean values
   d. _parse_date()            — str → datetime
   e. resolve_ou()             — account_number → ou_number + business_unit
   f. append NormalizedCreditRow
"""
from __future__ import annotations

import datetime as dt
import logging
import os
import re
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from ..common.regex_safety import safe_search
from .detector import DetectionResult
from .extractor import ExtractorFactory
from .credit_rules import eval_credit_rule
from .transforms import apply_transforms
from .ou_resolver import resolve_ou
from .currency import normalize_currency
from .account_locator import split_accounts, normalize_account, match_key

logger = logging.getLogger("cashapply.parser")


# ---------------------------------------------------------------------------
# Public output schema  (unchanged from previous version)
# ---------------------------------------------------------------------------

@dataclass
class NormalizedCreditRow:
    """
    One credit row after parsing and normalization.
    Produced by this module; consumed by extraction/chunk_processor.py
    via schemas.CreditRowSchema.
    """
    bank_name:      str
    account_number: str
    currency:       str
    narrative:      str
    credit_amount:  float
    statement_date: Optional[dt.datetime]
    bank_reference: Optional[str]
    ou_number:      Optional[str]
    business_unit:  Optional[str]


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class ColumnValidationError(Exception):
    pass


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------

_FORMAT_MAP = {
    "YYYY-MM-DD":  "%Y-%m-%d",
    "DD/MM/YYYY":  "%d/%m/%Y",
    "MM/DD/YYYY":  "%m/%d/%Y",
    "DD-MM-YYYY":  "%d-%m-%Y",
    "DD-Mon-YYYY": "%d-%b-%Y",
    "DD-MMM-YYYY": "%d-%b-%Y",
    # Month-name with a 2-DIGIT year (e.g. "4-Aug-26"). %Y needs a 4-digit
    # year, so the *-YYYY forms above FAIL on a 2-digit year — these %y forms
    # are what actually parse it. %d/%b are lenient about a single-digit day
    # ("4") and 3-letter month ("Aug"), so no separate single-digit variant is
    # needed. Month-name formats never match a purely numeric date, so adding
    # them can't cannibalise the numeric formats above.
    "DD-Mon-YY":   "%d-%b-%y",
    "DD-MMM-YY":   "%d-%b-%y",
    # 2-digit-year numeric variants — appended so the 4-digit forms above are
    # always tried first (a 4-digit year only matches %Y, a 2-digit only %y,
    # so the two never collide). Same day-first order as the 4-digit forms
    # (see ConfigBuilderWizard.tsx DEFAULT_DATE_FORMATS — numeric ordering
    # deliberately left day-first for now).
    "DD/MM/YY":    "%d/%m/%y",
    "MM/DD/YY":    "%m/%d/%y",
    "DD-MM-YY":    "%d-%m-%y",
    # ── Extra keys the config-builder date-format DETECTOR can resolve to (see
    # bank_statement/date_inference.py). Kept here so a resolved format is
    # stored as a readable key and parsed by the SAME map at runtime. Month-
    # first (MM-DD) and dot-separated and space/comma month-name variants live
    # here rather than in the wizard's blind default list — detection picks the
    # exact one from the sample, so they never widen runtime ambiguity.
    "YYYY/MM/DD":   "%Y/%m/%d",
    "MM-DD-YYYY":   "%m-%d-%Y",
    "MM-DD-YY":     "%m-%d-%y",
    "DD.MM.YYYY":   "%d.%m.%Y",
    "MM.DD.YYYY":   "%m.%d.%Y",
    "DD.MM.YY":     "%d.%m.%y",
    "MM.DD.YY":     "%m.%d.%y",
    "DD Mon YYYY":  "%d %b %Y",
    "DD Mon YY":    "%d %b %y",
    "Mon DD YYYY":  "%b %d %Y",
    "Mon DD, YYYY": "%b %d, %Y",
    "Mon-DD-YYYY":  "%b-%d-%Y",
}


def _excel_serial_to_date(n: float) -> Optional[dt.datetime]:
    """Convert an Excel/1900 date serial number to a datetime.

    Excel day 1 = 1900-01-01, with the well-known phantom 1900-02-29 leap-year
    bug; the standard correction is an epoch of 1899-12-30. Only plausible date
    serials (≈ 1970-2100) are accepted so genuine numeric values aren't mistaken
    for dates.
    """
    if not (20000 <= n <= 80000):
        return None
    try:
        return dt.datetime(1899, 12, 30) + dt.timedelta(days=int(n))
    except (OverflowError, ValueError):
        return None


def strip_time_component(s: str) -> str:
    """
    "2026-01-06 00:00:00" / "2026-01-06T00:00:00" -> "2026-01-06"; anything else
    unchanged.

    Datetime-looking strings are extremely common in practice: pandas renders a
    real Excel/CSV date cell as a Timestamp, and str() of that always carries a
    00:00:00 tail. A date FORMAT like %Y-%m-%d cannot parse that tail, so the
    tail has to come off before the format is tried.

    PUBLIC, and shared with date_inference.py on purpose. That module decides
    which format to store in a recipe by testing candidates against real sample
    values, so if it were stricter than this parser it would report "no known
    format parses these" for values the parser reads perfectly -- which is
    exactly the bug this helper was extracted to fix. Keep the two using one
    implementation.

    The guard matters: the pattern requires a time-looking token after the
    separator, so a value that merely CONTAINS a space (never a valid date
    format here, but cheap to be safe about) is not truncated blindly.
    """
    return re.split(r"[ T]", s, maxsplit=1)[0] if re.search(r"[ T]\d{1,2}:", s) else s


def _parse_date(value, date_formats: list) -> Optional[dt.datetime]:
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    # Excel serial date (number or numeric string with no separators)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return _excel_serial_to_date(float(value))

    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return None

    # Trim a trailing time component ("2026-01-06 00:00:00" / "...T00:00:00")
    s_core = strip_time_component(s)

    for fmt_key in date_formats:
        for candidate in (s, s_core):
            try:
                return dt.datetime.strptime(candidate, _FORMAT_MAP.get(fmt_key, fmt_key))
            except ValueError:
                continue

    # Bare numeric string → Excel serial fallback
    if re.fullmatch(r"\d+(\.\d+)?", s):
        return _excel_serial_to_date(float(s))

    return None


# ---------------------------------------------------------------------------
# Amount parsing  (handles thousands separators, parentheses/sign, CR/DR
# suffixes, currency symbols, and European decimal notation)
# ---------------------------------------------------------------------------

def parse_amount(value) -> Optional[float]:
    """Parse a monetary value into a float, or None if it isn't a number.

    Handles the formats real bank statements use:
      "1,234.56"  "1.234,56"  "(1,234.56)"  "1234.56 CR"  "$ 1,234.56"  "-50"
    Returns None (not 0.0) when the value is genuinely non-numeric so callers
    can distinguish "no amount" from "zero".
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return None

    negative = False

    # Parentheses → negative, e.g. (1,234.56)
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    # Trailing/leading CR/DR markers (DR = debit → negative)
    m = re.search(r"\b(CR|DR)\b\.?$", s, re.IGNORECASE)
    if m:
        if m.group(1).upper() == "DR":
            negative = True
        s = s[:m.start()].strip()

    # Strip currency symbols / letters / spaces, keep digits, separators, sign
    s = re.sub(r"[^\d.,\-+]", "", s)
    if not s or s in ("-", "+", ".", ","):
        return None

    if s.startswith("-"):
        negative = True
    s = s.lstrip("+-")

    # Decide decimal separator: the last-occurring of '.' or ',' is the decimal
    last_dot, last_comma = s.rfind("."), s.rfind(",")
    if last_dot == -1 and last_comma == -1:
        cleaned = s
    elif last_comma > last_dot:
        # European: ',' is decimal, '.' is thousands
        cleaned = s.replace(".", "").replace(",", ".")
    else:
        # Anglo: '.' is decimal, ',' is thousands
        cleaned = s.replace(",", "")

    try:
        amount = float(cleaned)
    except ValueError:
        return None
    return -amount if negative else amount


# ---------------------------------------------------------------------------
# Column validation
# ---------------------------------------------------------------------------

def _validate_columns(df: pd.DataFrame, fields: list, credit_rule: dict | None = None) -> None:
    required: list[str] = []
    for f in fields:
        src = f["from"]
        if src["type"] == "column" and src.get("name"):
            required.append(src["name"])
        elif src["type"] == "concat":
            required.extend(src.get("names", []))

    # The credit rule's flag/amount column must also exist, otherwise the rule
    # silently matches no rows and the run produces "0 credit rows" with no clue.
    if credit_rule:
        rule_col = _credit_rule_column(credit_rule, fields)
        if rule_col and rule_col not in required:
            required.append(rule_col)

    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ColumnValidationError(
            f"Config expects columns {missing} but they were not found in the file. "
            f"Available columns: {list(df.columns)}"
        )


def _credit_rule_column(credit_rule: dict, fields: list) -> str | None:
    """Resolve the actual DataFrame column the credit rule reads.

    For amount-based rules the rule.field is a *logical* field name that maps to
    a column; for flag_matches the rule.field is the raw column name directly.
    """
    rule_type = credit_rule.get("type")
    field = credit_rule.get("field")
    if not field:
        return None
    if rule_type in ("amount_positive", "column_not_blank"):
        for f in fields:
            if f["name"] == field and f.get("from", {}).get("type") == "column":
                return f["from"].get("name")
        return field
    return field  # flag_matches → raw column name


def _row_account(value, registered: str = "") -> str:
    """The account a row's receipt is posted against.

    Normalizes with account_locator.normalize_account — the SAME implementation
    detection and the registry use. (This was a local normalizer that only
    trimmed a trailing ".0", so spaces/dashes survived here but not there, and
    the account stored on a row could fail to match the BankAccount it meant.)

    `registered` is the account this config is registered under. It wins whenever
    the cell names several accounts — a header cell reading
    "41678876 & 41678884" resolves to the registered account, not to whichever
    token happens to come first. That's what makes a cell/fixed-mapped statement
    deterministic: one config, one account, every row.
    """
    accounts = split_accounts(value)
    if not accounts:
        return normalize_account(registered)
    if len(accounts) > 1:
        reg = normalize_account(registered)
        if reg and match_key(reg) in {match_key(a) for a in accounts}:
            return reg
        # No registered account to fall back on (a wizard test parse) — take the
        # first token rather than inventing a joined value that matches nothing.
        logger.warning(
            "Account cell %r names %d accounts and none is the registered account "
            "(%r) — using %r. Register this config under one of them.",
            value, len(accounts), registered, accounts[0],
        )
    return accounts[0]


# ---------------------------------------------------------------------------
# Cell reader  (used for cell-type fields — reads once before the row loop)
# ---------------------------------------------------------------------------

def _read_cell(filepath: str, cfg: dict, row: int, col: int) -> str:
    """Read a single cell from the file (metadata position, 0-indexed).

    Handles .csv/.txt, .xls and .xlsx. The CSV branch is NOT optional garnish:
    a `cell`-type field is how a recipe picks up statement metadata that sits
    ABOVE the table -- the account number, currency or bank name in a preamble
    block -- and that layout is just as common in a .csv export as in a
    spreadsheet.

    Before this branch existed, everything that was not .xls fell through to
    openpyxl's load_workbook(), so pointing a cell field at a CSV failed with
    "openpyxl does not support .csv file format", naming a library the user
    never chose and giving no hint that the real problem was the file format.
    """
    source = cfg.get("source", {})
    sheet_cfg = source.get("sheet", {})

    if filepath.lower().endswith((".csv", ".txt")):
        # No sheets in a CSV -- sheet_cfg is meaningless here, and the position
        # is simply the Nth field of the Nth line. Delimiter and encoding are
        # resolved exactly as the extractor resolves them, so a cell field and
        # a column field can never disagree about where the columns are.
        import csv as _csv

        from .extractor.csv_extractor import resolve_encodings, sniff_dialect

        declared_delim = source.get("delimiter", "auto")
        for enc in resolve_encodings(source.get("encoding", "auto")):
            try:
                delim = (
                    sniff_dialect(filepath, enc) if declared_delim == "auto"
                    else declared_delim
                )
                with open(filepath, encoding=enc, newline="") as f:
                    for r, values in enumerate(_csv.reader(f, delimiter=delim)):
                        if r == row:
                            return values[col].strip() if col < len(values) else ""
                return ""   # requested row is past the end of the file
            except UnicodeDecodeError:
                continue    # wrong encoding -- try the next candidate
        return ""

    if filepath.lower().endswith(".xls"):
        try:
            import xlrd
        except ImportError:
            return ""
        wb = xlrd.open_workbook(filepath)
        by    = sheet_cfg.get("by", "first")
        value = sheet_cfg.get("value")
        if by == "name" and value and value in wb.sheet_names():
            ws = wb.sheet_by_name(value)
        else:
            ws = wb.sheet_by_index(0)
        try:
            val = ws.cell_value(row, col)
        except IndexError:
            return ""
        return str(val).strip() if val not in (None, "") else ""

    wb = load_workbook(filepath, read_only=True, data_only=True)
    by    = sheet_cfg.get("by", "first")
    value = sheet_cfg.get("value")
    if by == "name" and value and value in wb.sheetnames:
        ws = wb[value]
    else:
        ws = wb[wb.sheetnames[0]]
    rows = list(
        ws.iter_rows(
            min_row=row + 1, max_row=row + 1,
            min_col=col + 1, max_col=col + 1,
            values_only=True,
        )
    )
    wb.close()
    val = rows[0][0] if rows and rows[0] else None
    return str(val).strip() if val is not None else ""


# ---------------------------------------------------------------------------
# File-level field resolver (cell / fixed / filename_pattern)
# ---------------------------------------------------------------------------

def _resolve_file_fields(filepath: str, fields: list, cfg: dict) -> dict:
    filename = os.path.basename(filepath)
    result: dict = {}
    for f in fields:
        name = f["name"]
        src  = f["from"]
        if src["type"] == "fixed":
            result[name] = src.get("value", "")
        elif src["type"] == "cell":
            result[name] = _read_cell(filepath, cfg, src["row"], src["col"])
        elif src["type"] == "filename_pattern":
            # safe_search() validates the config-supplied pattern (length,
            # nested quantifiers, syntax) and bounds the scanned text before
            # it reaches the backtracking engine -- see common/regex_safety.py.
            m = safe_search(src.get("pattern", ""), filename)
            group = src.get("group", 1)
            result[name] = m.group(group) if m else ""
    return result


# ---------------------------------------------------------------------------
# Row-level field resolver (column / concat)
# ---------------------------------------------------------------------------

def _resolve_row_field(row: pd.Series, src: dict) -> str:
    if src["type"] == "column":
        name = src.get("name")
        if name is None:   # null in JSON → no column configured
            return ""
        val = row.get(name, "")
        # Guard: duplicate column names cause row.get() to return a Series
        if isinstance(val, pd.Series):
            val = val.iloc[0] if len(val) > 0 else ""
        return str(val if val else "").strip()
    if src["type"] == "concat":
        sep   = src.get("sep", " ")
        parts = [str(row.get(n, "") or "") for n in src.get("names", [])]
        return sep.join(parts).strip()
    return ""


# ---------------------------------------------------------------------------
# Exclusion filter
# ---------------------------------------------------------------------------

def _apply_exclusions(df: pd.DataFrame, exclusions: list) -> pd.DataFrame:
    for exc in exclusions:
        t     = exc["type"]
        field = exc.get("field", "")

        if t == "field_value_in":
            values_lower = [v.lower() for v in exc.get("values", [])]
            if field in df.columns:
                df = df[~df[field].astype(str).str.lower().isin(values_lower)]

        elif t == "field_not_equals":
            if field in df.columns:
                df = df[df[field].astype(str).str.strip() == str(exc.get("value", ""))]

        elif t == "field_blank":
            if field in df.columns:
                df = df[
                    df[field].notna()
                    & (df[field].astype(str).str.strip() != "")
                ]

        # NOTE: a "field_matches" exclusion type used to live here, applying a
        # user-typed regex to EVERY row of the statement -- a ReDoS sink
        # (CWE-1333) whose cost scaled with row count. It has been removed
        # from both the wizard and this parser; the remaining three types
        # cover the documented use case (skipping Opening/Closing Balance
        # rows). An old recipe still carrying one is ignored rather than
        # honoured -- falling through here drops no rows, which is the safe
        # direction: an unfiltered row surfaces for review, it is not
        # silently swallowed.
    return df


# ---------------------------------------------------------------------------
# Public parse function
# ---------------------------------------------------------------------------

def parse_credit_rows(
    filepath: str,
    detection: DetectionResult,
    filename: str,
) -> list[NormalizedCreditRow]:
    """
    Parse the file at `filepath` using the detected bank config.
    Returns only credit rows (as defined by the config's credit_rule).
    """
    cfg = detection.config
    if cfg is None:
        raise ValueError(
            f"No config resolved for {filename}; manual assignment required."
        )

    fields = cfg.get("fields", [])

    # 1. Extract DataFrame. Identifier columns are read as TEXT: pandas infers a
    # column of digit strings as int64 (or float64 if any row is blank), so an
    # account stored as "000274178" would arrive as 274178 — the receipt would be
    # created against the wrong account number, and a 16+ digit account would lose
    # precision outright. Same reasoning for the bank reference, which feeds the
    # row hash and Oracle matching. Amounts and dates are excluded on purpose;
    # they rely on pandas' numeric/datetime handling.
    _IDENTIFIER_FIELDS = ("account_number", "bank_reference")
    text_columns: list[str] = []
    for f in fields:
        if f.get("name") not in _IDENTIFIER_FIELDS:
            continue
        src = f.get("from", {})
        if src.get("type") == "column" and src.get("name"):
            text_columns.append(src["name"])
        elif src.get("type") == "concat":
            text_columns.extend(src.get("names", []))
    df = ExtractorFactory.extract(filepath, cfg["source"], text_columns=text_columns or None)

    # 2. Validate required columns exist — loud error if not
    _validate_columns(df, fields, cfg.get("credit_rule"))

    # 3. Resolve file-level fields (cell, fixed, filename_pattern) — once
    file_fields = _resolve_file_fields(filepath, fields, cfg)

    # 4. Apply row exclusions before the credit rule
    df = _apply_exclusions(df, cfg.get("exclusions", []))

    # 5. Iterate rows
    credit_rule  = cfg["credit_rule"]
    date_formats = cfg.get("date_formats", [])
    transforms   = cfg.get("transforms", {})
    # Fall back to the human-readable display name (never the internal config key)
    default_bank = cfg.get("display_name") or ""
    # ISO fallback stamped into the recipe at save (config_builder_routes.builder_save)
    # — used when a row's own currency value can't be standardized to ISO.
    config_currency = normalize_currency(cfg.get("currency")) if cfg.get("currency") else None
    # The account this config is registered under (stamped by detector's
    # _recipe_config). Governs the row account whenever the mapped cell names
    # several accounts — see _row_account. Absent on a wizard test parse.
    registered_account = str(cfg.get("account_number") or "")

    out: list[NormalizedCreditRow] = []
    invalid_amount_count = 0
    unmapped_currency_count = 0

    for _, row in df.iterrows():
        # 5a. Credit rule
        if not eval_credit_rule(row, credit_rule, fields):
            continue

        # 5b. Resolve per-row fields (column, concat)
        record: dict = dict(file_fields)
        for f in fields:
            src = f["from"]
            if src["type"] in ("column", "concat"):
                record[f["name"]] = _resolve_row_field(row, src)

        # 5c. Apply transforms
        record = apply_transforms(record, transforms)

        # 5d. Parse date
        statement_date = _parse_date(record.get("date"), date_formats)

        # 5e. Convert credit_amount to float (robust: commas, parens, CR/DR, €)
        credit_amount = parse_amount(record.get("credit_amount"))
        if credit_amount is None or credit_amount <= 0:
            # A Credit Amount must be numeric and strictly POSITIVE (> 0)
            # (TC-180). A credit-rule-accepted row whose amount is non-numeric,
            # negative, or zero is INVALID — flag it (WARNING log, with context)
            # and DROP it, never ingest it as a valid credit. A zero-value
            # "credit" isn't an applicable payment, and negatives/zeros matter
            # especially for flag-based credit rules (flag_matches), which only
            # look at a CR/DR marker and never at the amount — so a negative or
            # zero "credit" would otherwise slip through here as a real credit.
            # This also keeps flag-based rules consistent with amount_positive,
            # which already requires `amount > 0`.
            reason = ("non-numeric" if credit_amount is None
                      else "zero" if credit_amount == 0 else "negative")
            invalid_amount_count += 1
            logger.warning(
                "Dropping row with %s credit amount %r (file=%s, ref=%s, date=%r)",
                reason, record.get("credit_amount"), filename,
                record.get("bank_reference"), record.get("date"),
            )
            continue

        # 5f. OU resolution — strictly per row's own account (account_ou_map).
        # No fallback to the detection account: in a multi-account file that would
        # wrongly stamp every row with the identifying account's OU.
        account_number = _row_account(record.get("account_number", ""), registered_account)
        ou_info = resolve_ou(account_number)
        ou_number     = ou_info.get("ou_number")
        business_unit = ou_info.get("business_unit")

        # 5g. Standardize currency to ISO-4217 (Fusion requires it). Map known
        # spellings ("EURO"->"EUR" etc.); if a row's value can't be mapped, fall
        # back to the config's currency and flag it (never blocks ingestion).
        raw_currency = record.get("currency", "")
        currency = normalize_currency(raw_currency)
        if currency is None:
            currency = config_currency or str(raw_currency or "").upper().strip()
            if raw_currency:
                unmapped_currency_count += 1
                logger.warning(
                    "Currency %r not a known ISO code (file=%s, ref=%s) — fell back to %r",
                    raw_currency, filename, record.get("bank_reference"), currency,
                )

        out.append(NormalizedCreditRow(
            bank_name      = record.get("bank_name") or default_bank,
            account_number = account_number,
            currency       = currency,
            narrative      = record.get("narrative", ""),
            credit_amount  = credit_amount,
            statement_date = statement_date,
            bank_reference = record.get("bank_reference") or None,
            ou_number      = ou_number,
            business_unit  = business_unit,
        ))

    if invalid_amount_count:
        logger.warning(
            "%s: dropped %d row(s) with invalid (non-numeric, zero, or negative) credit amounts",
            filename, invalid_amount_count,
        )
    if unmapped_currency_count:
        logger.warning(
            "%s: %d row(s) had a currency that couldn't be mapped to an ISO code — fell back to the config currency (%s)",
            filename, unmapped_currency_count, config_currency or "raw value",
        )

    return out
