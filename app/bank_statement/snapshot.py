"""
app.bank_statement.snapshot
===========================
FileSnapshot — read a file once and cache raw cell values for cheap lookups.
Used by the account locator (and the account-based detector) to read header
cells / scan the top region without re-opening the file per check.

Standalone (no dependency on detector) so the detector can import it without
circular imports.
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field

_MAX_ROWS = 50
_MAX_COLS = 25


@dataclass
class FileSnapshot:
    filepath:    str
    filename:    str
    extension:   str                       # "xlsx"/"xls"/"csv"/… (lowercase, no dot)
    sheet_names: list[str] = field(default_factory=list)
    # cells keyed by (sheet_name_lower, row_0idx, col_0idx) -> str
    cells:       dict = field(default_factory=dict)

    @classmethod
    def from_path(cls, filepath: str) -> "FileSnapshot":
        filename = os.path.basename(filepath)
        ext = os.path.splitext(filename)[1].lower().lstrip(".")
        sheets: list[str] = []
        cells: dict = {}

        if ext in ("xlsx", "xlsm"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath, read_only=True, data_only=True)
                sheets = list(wb.sheetnames)
                for sh in sheets:
                    ws = wb[sh]
                    key = sh.lower()
                    for r, row in enumerate(
                        ws.iter_rows(min_row=1, max_row=_MAX_ROWS, max_col=_MAX_COLS, values_only=True)
                    ):
                        for c, val in enumerate(row):
                            cells[(key, r, c)] = str(val).strip() if val is not None else ""
                wb.close()
            except Exception:
                pass

        elif ext == "xls":
            try:
                import xlrd
                wb = xlrd.open_workbook(filepath)
                sheets = wb.sheet_names()
                for sh in sheets:
                    ws = wb.sheet_by_name(sh)
                    key = sh.lower()
                    for r in range(min(_MAX_ROWS, ws.nrows)):
                        for c in range(min(_MAX_COLS, ws.ncols)):
                            val = ws.cell_value(r, c)
                            cells[(key, r, c)] = str(val).strip() if val not in (None, "") else ""
            except Exception:
                pass

        elif ext in ("csv", "txt"):
            # Delimiter and encoding come from the same helpers the extractor
            # uses (extractor/csv_extractor.py), so the snapshot sees the file
            # split exactly the way extraction will.
            #
            # This used to sniff inline and trust the result. csv.Sniffer can
            # return "\r" on a ragged header block with CRLF line endings, which
            # csv.reader then rejects -- and because this branch swallows
            # exceptions, that produced an EMPTY snapshot instead of an error.
            # No cells means the account locator finds nothing, so the file
            # simply looks unrecognised, for a reason pointing nowhere near the
            # delimiter. sniff_dialect() validates the result, and the encoding
            # is now tried properly rather than forced through utf-8 with
            # errors="replace" (which turned latin-1 names into mojibake).
            import csv as _csv

            from .extractor.csv_extractor import resolve_encodings, sniff_dialect

            for enc in resolve_encodings("auto"):
                try:
                    delim = sniff_dialect(filepath, enc)
                    with open(filepath, encoding=enc, newline="") as f:
                        for r, row in enumerate(_csv.reader(f, delimiter=delim)):
                            if r >= _MAX_ROWS:
                                break
                            for c, val in enumerate(row[:_MAX_COLS]):
                                cells[("sheet1", r, c)] = str(val).strip()
                    sheets = ["Sheet1"]
                    break
                except UnicodeDecodeError:
                    # Wrong encoding, and the read may have stopped part-way --
                    # discard what it collected before retrying, or the snapshot
                    # would mix rows from two decodings.
                    cells.clear()
                    continue
                except Exception:
                    cells.clear()
                    break

        return cls(filepath=filepath, filename=filename, extension=ext,
                   sheet_names=sheets, cells=cells)

    def cell(self, row: int, col: int, sheet: str | None = None) -> str:
        """Value at (row,col). If sheet given, read that sheet; else first non-empty
        across all sheets."""
        if sheet is not None:
            return self.cells.get((sheet.lower(), row, col), "")
        for sh in self.sheet_names:
            v = self.cells.get((sh.lower(), row, col), "")
            if v:
                return v
        return ""

    def iter_values(self, sheet: str | None = None):
        """Yield every cached cell string (optionally for one sheet)."""
        for (sh, _r, _c), v in self.cells.items():
            if v and (sheet is None or sh == sheet.lower()):
                yield v
