"""
app.bff.config_builder_routes — /api/config/*  (Account-Based Ingestion)
========================================================================
Config Builder + account-config management for the account-based engine.

  GET    /builder/raw-preview/{filename}   — raw cell grid for the wizard
  POST   /builder/locate-account           — live-extract account(s) via a locator draft
  POST   /builder/test                     — test a *draft recipe* against a file
  GET    /builder/available-ous            — OU/BU picklist for the wizard's OU step
  POST   /builder/save                     — save a recipe (create account / add format recipe)
  GET    /builder/accounts                 — list accounts (Manage + Clone-from-existing)
  GET    /builder/account/{account_number} — full account entry (for cloning)
  DELETE /builder/{account_number}          — delete an account (all its recipes)
  DELETE /builder/{account_number}/{format} — delete a single format recipe
  POST   /test-existing                    — test a *saved* recipe against a file
  GET    /detect/{filename}                — re-detect + list matching accounts

DB-BACKED (was JSON only). account_configs.json, bank_ou_mapping.json,
account_ou_map.json and ou_functional_currency.json are no longer read or
written anywhere in this module. OU + Business Unit are captured here at
save time as a real relationship — BankAccount.ou_id -> OrganizationUnit —
via _get_or_create_organization_unit()/_get_or_create_bank_account() below,
and recipes are versioned rows in AccountConfigRecipe. This is now the
single place account configs are created; everything downstream (detector,
parser, ou_resolver, fx_service) reads the same DB rows through
bank_statement/configs/account_loader.py.
Register under /api/config alongside config_routes.
"""
from __future__ import annotations

import os
import logging
import dataclasses
import datetime as dt

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..db.models import SourceFile, User, BankAccount, OrganizationUnit, AccountConfigRecipe
from ..deps import get_db
from ..auth import require_permission
from ..audit.service import log_activity
from ..storage.client import get_storage_client
from ..common.errors import AppError
from ..common.error_codes import ErrorCode
from ..common.upload_validation import validate_statement_upload, validate_statement_size
from ..bank_statement.account_locator import (
    extract_accounts, normalize_account, last4, match_key,
)
from ..bank_statement.account_validation import account_reject_reason
from ..bank_statement.field_sanity import check_field_values
from ..bank_statement.currency import normalize_currency
from ..bank_statement.configs.account_loader import (
    load_account_configs, load_account_ou_map, reload_account_configs,
    active_recipe, format_summaries,
)
from ..aging import aging_store

router = APIRouter()
logger = logging.getLogger("cashapply.config_builder")

_STATEMENT_BUCKET = "bank-statements"
_FMT_ALIASES = {"xlsm": "xlsx", "txt": "csv"}


# ── shared helpers ────────────────────────────────────────────────────────────

def _local_path(db: Session, filename: str) -> tuple[SourceFile, str]:
    record = db.query(SourceFile).filter(
        SourceFile.kind == "bank_statement", SourceFile.filename == filename
    ).first()
    if not record:
        raise AppError(ErrorCode.STATEMENT_NOT_FOUND, detail=f"file '{filename}'")
    storage = get_storage_client()
    return record, storage.local_path_for_read(_STATEMENT_BUCKET, record.storage_key)


def _local_path_by_key(db: Session, storage_key: str) -> tuple[SourceFile, str]:
    record = db.query(SourceFile).filter(
        SourceFile.kind == "bank_statement", SourceFile.storage_key == storage_key
    ).first()
    if not record:
        raise AppError(ErrorCode.STATEMENT_NOT_FOUND)
    storage = get_storage_client()
    return record, storage.local_path_for_read(_STATEMENT_BUCKET, record.storage_key)


def _file_format(filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower().lstrip(".")
    return _FMT_ALIASES.get(ext, ext)


def _test_recipe(local_path: str, recipe: dict, key: str, filename: str) -> dict:
    """Parse a file with a recipe and return a row-count + sample preview."""
    from ..bank_statement.detector import DetectionResult
    from ..bank_statement.parser import parse_credit_rows, ColumnValidationError

    detection = DetectionResult(success=True, config_key=key, config=recipe,
                                step_used=1, method_detail=f"Test '{key}'")
    try:
        rows = parse_credit_rows(local_path, detection, filename)

        def _ser(r):
            d = dataclasses.asdict(r)
            sd = d.get("statement_date")
            if sd is not None:
                # statement_date is a datetime (bank dates carry no time, so it's
                # always midnight) — show the date only, so the preview doesn't
                # display a phantom "00:00:00". Fall back to str() for anything
                # unexpected that lacks .date().
                d["statement_date"] = sd.date().isoformat() if hasattr(sd, "date") else str(sd)
            return d

        # Value-level sanity checks over the REAL parsed values (account/date/
        # currency/narrative). account_number failures are severity "error";
        # everything else is "warn". `account_ok` is the single flag the wizard
        # gates on — a config can parse cleanly yet still be wired to the wrong
        # columns (e.g. a metadata label cell as the account number), which is
        # exactly the "passed the test but was wrong" hole this closes.
        warnings = check_field_values(rows)
        account_ok = not any(
            w["field"] == "account_number" and w["severity"] == "error" for w in warnings
        )
        return {
            "success": True,
            "row_count": len(rows),
            "rows": [_ser(r) for r in rows[:50]],
            "warnings": warnings,
            "account_ok": account_ok,
        }
    except (ColumnValidationError, Exception) as e:
        return {"success": False, "error": str(e), "row_count": 0, "rows": [],
                "warnings": [], "account_ok": False}


# ── upload a file for the wizard (NO ingestion) ──────────────────────────────────

@router.post("/builder/upload")
async def builder_upload(file: UploadFile = File(...), db: Session = Depends(get_db),
                          user: User = Depends(require_permission("config:manage"))):
    """
    Store an uploaded report so the Config Builder wizard can preview / locate /
    test against it — WITHOUT running the ingestion pipeline.

    Config-building is by definition the "no config exists yet" case, so routing
    this through /api/run/upload (which detect_config()s and then defers
    ingest_statement) always fails detection ("Bank format not auto-detected"),
    marks the file ingest_status="error", and clutters the Home analysis list
    with a failed file. Here we only need the bytes on disk + a SourceFile row so
    raw-preview can resolve the file by name.

    - No ingestion is deferred.
    - No duplicate-hash block: re-uploading the same file (to add a NEW version
      of an existing config) is expected and must be allowed.
    - The row is created archived=True so it never shows in the Home statements
      list; the wizard reads it by (kind, filename) regardless of archived state.
    """
    # Only Excel/CSV, 10 MB max — reject anything else with a clear message
    # before reading the bytes or touching storage.
    validate_statement_upload(file.filename)
    validate_statement_size(file.size)
    filename = file.filename or "upload"
    data = await file.read()
    validate_statement_size(len(data))

    storage = get_storage_client()
    storage.save(_STATEMENT_BUCKET, filename, data)

    record = db.query(SourceFile).filter(
        SourceFile.kind == "bank_statement", SourceFile.filename == filename
    ).first()
    if record:
        # Reuse the existing row (e.g. adding a new version of the same file).
        record.storage_key = filename
    else:
        record = SourceFile(
            kind="bank_statement",
            filename=filename,
            storage_key=filename,
            ingest_status="config_only",   # never analysed; excluded from Home list
            archived=True,
        )
        db.add(record)
    db.commit()
    db.refresh(record)

    return {"filename": filename, "source_file_id": record.id}


# ── raw preview ─────────────────────────────────────────────────────────────────

def _trim_trailing_empty_cols(rows: list[list[str]]) -> list[list[str]]:
    """Normalize every row to equal width, then drop only the TRAILING columns
    that are empty in every previewed row. This is what makes the wizard show
    exactly the columns that carry data — a column is kept if it has a value in
    the header OR any data row, so nothing with data is missed, while the run of
    blank padding columns after the last real one is removed. Interior empty
    columns are preserved (index position matters for column mapping)."""
    if not rows:
        return rows
    width = max((len(r) for r in rows), default=0)
    norm = [list(r) + [""] * (width - len(r)) for r in rows]
    last_with_data = -1
    for c in range(width):
        if any((norm[r][c] or "").strip() for r in range(len(norm))):
            last_with_data = c
    keep = last_with_data + 1
    return [r[:keep] for r in norm]


@router.get("/builder/raw-preview/{filename}")
def builder_raw_preview(filename: str, db: Session = Depends(get_db),
                        user: User = Depends(require_permission("config:manage"))):
    """Return raw cell data for all sheets — used by the wizard preview/header/locate steps."""
    record, local_path = _local_path(db, filename)
    ext = os.path.splitext(filename)[1].lower().lstrip(".")

    # MAX_COLS is a safety CEILING against pathologically wide files, not the
    # display width — actual columns returned are trimmed to the last one that
    # holds data (see _trim_trailing_empty_cols), so a normal statement shows
    # all its real columns (e.g. 27) without hitting the ceiling.
    MAX_ROWS, MAX_COLS = 40, 50
    sheets = []
    if ext in ("xlsx", "xlsm"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(local_path, read_only=True, data_only=True)
            for sh in wb.sheetnames:
                ws = wb[sh]
                rows = [[str(v).strip() if v is not None else "" for v in row]
                        for row in ws.iter_rows(min_row=1, max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True)]
                sheets.append({"name": sh, "rows": _trim_trailing_empty_cols(rows)})
            wb.close()
        except Exception as e:
            logger.exception("raw-preview: failed to read xlsx %s", filename)
            raise AppError(ErrorCode.CONFIG_FILE_UNREADABLE,
                            detail=f"'{filename}' doesn't look like a valid Excel (.xlsx) file ({e.__class__.__name__})")
    elif ext == "xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(local_path)
            for sh in wb.sheet_names():
                ws = wb.sheet_by_name(sh)
                rows = []
                for r in range(min(MAX_ROWS, ws.nrows)):
                    row = [str(ws.cell_value(r, c)).strip() if ws.cell_value(r, c) not in (None, "") else ""
                           for c in range(min(MAX_COLS, ws.ncols))]
                    row += [""] * (MAX_COLS - len(row))
                    rows.append(row)
                sheets.append({"name": sh, "rows": _trim_trailing_empty_cols(rows)})
        except Exception as e:
            logger.exception("raw-preview: failed to read xls %s", filename)
            raise AppError(ErrorCode.CONFIG_FILE_UNREADABLE,
                            detail=f"'{filename}' doesn't look like a valid legacy Excel (.xls) file ({e.__class__.__name__})")
    elif ext in ("csv", "txt"):
        # This preview MUST split the file exactly as extraction will, so it
        # reuses the extractor's own delimiter sniffing and encoding fallback
        # (bank_statement/extractor/csv_extractor.py) instead of assuming
        # comma + utf-8.
        #
        # Both were hardcoded here previously, which meant a semicolon-delimited
        # or latin-1 statement rendered as ONE mangled column in the wizard grid
        # while extracting perfectly well. That is not merely cosmetic: the
        # wizard asks the user to click a row to mark the header (and now the
        # sub-header), and neither can be picked out of a single joined column.
        import csv
        from ..bank_statement.extractor.csv_extractor import resolve_encodings, sniff_dialect

        rows = None
        for enc in resolve_encodings("auto"):
            try:
                delim = sniff_dialect(local_path, enc)
                with open(local_path, encoding=enc, newline="") as f:
                    rows = [[str(v).strip() for v in row[:MAX_COLS]]
                            for i, row in enumerate(csv.reader(f, delimiter=delim))
                            if i < MAX_ROWS]
                break
            except UnicodeDecodeError:
                continue   # wrong encoding for this file -- try the next candidate
            except Exception as e:
                logger.exception("raw-preview: failed to read csv %s", filename)
                raise AppError(ErrorCode.CONFIG_FILE_UNREADABLE,
                                detail=f"'{filename}' could not be parsed as CSV ({e.__class__.__name__})")

        if rows is None:
            logger.error("raw-preview: could not decode csv %s with any known encoding", filename)
            raise AppError(ErrorCode.CONFIG_FILE_UNREADABLE,
                            detail=f"'{filename}' could not be read as text (tried UTF-8 and Latin-1)")

        sheets.append({"name": "Sheet1", "rows": _trim_trailing_empty_cols(rows)})
    else:
        raise AppError(ErrorCode.CONFIG_FILE_TYPE_UNSUPPORTED, detail=f"'.{ext}' extension")

    return {"filename": filename, "storage_key": record.storage_key, "extension": ext, "sheets": sheets}


# ── locate account (live preview for the wizard's account step) ──────────────────

class LocateAccountRequest(BaseModel):
    storage_key: str
    locator: dict
    source: dict | None = None
    # True when the recipe's per-row account_number field is a COLUMN, i.e. rows
    # span accounts and every one of them needs its own config. Drives whether
    # this response invites a fan-out or a single "pick the account this config is
    # for" (see rows_span_accounts in bank_statement/detector.py).
    rows_span_accounts: bool = False


# How many discovered accounts the wizard is asked to render/onboard at once.
# A pathological column (or a locator pointed at the wrong column) can yield
# hundreds; truncate EXPLICITLY and say so rather than silently showing a subset.
_MAX_DISCOVERED_ACCOUNTS = 50


@router.post("/builder/locate-account")
def builder_locate_account(body: LocateAccountRequest, db: Session = Depends(get_db),
                           user: User = Depends(require_permission("config:manage"))):
    """Run a locator draft against the file and return the account(s) it finds.

    When the recipe's per-row account field is a COLUMN, rows span accounts and
    the wizard onboards every one of them against the same recipe (see
    builder_save's `accounts`), so this response carries what that table needs:
    which accounts are already configured, which don't look like real accounts,
    and any OU already on record to prefill with.

    Otherwise every row shares one account and the wizard asks the user to pick
    the single account this config is for — a header cell naming a main and its
    sub-account is a choice, not two configs to create.
    """
    _record, local_path = _local_path_by_key(db, body.storage_key)
    source = body.source or {}

    found_all = sorted(extract_accounts(local_path, body.locator, source))
    truncated = max(0, len(found_all) - _MAX_DISCOVERED_ACCOUNTS)
    found = found_all[:_MAX_DISCOVERED_ACCOUNTS]

    configs = load_account_configs()
    by_key = {match_key(entry.get("account_number", k)): (k, entry) for k, entry in configs.items()}

    existing: dict[str, list[str]] = {}
    known: dict[str, dict] = {}
    for acct in found:
        hit = by_key.get(match_key(acct))
        if not hit:
            continue
        _k, entry = hit
        existing[acct] = sorted((entry.get("recipes") or {}).keys())
        # Prefill for the Save step's per-account OU table, so onboarding N
        # accounts is mostly confirming rather than retyping.
        known[acct] = {
            "display_name":  entry.get("display_name"),
            "bank":          entry.get("bank"),
            "currency":      entry.get("currency"),
            "ou_number":     entry.get("ou_number"),
            "business_unit": entry.get("business_unit"),
        }

    # Flag candidates that don't look like real account numbers (e.g. the locator
    # landed on a label/heading cell, or a "TOTAL" footer row got scanned) so the
    # wizard can grey them out BEFORE one is chosen as an account identity.
    # {account: reason-or-None}.
    account_issues = {a: account_reject_reason(a) for a in found}
    return {
        "accounts": found,
        "count": len(found),
        "total_found": len(found_all),
        "truncated": truncated,             # >0 ⇒ tell the user N were not shown
        "last4s": sorted({last4(a) for a in found}),
        "existing": existing,               # {account: [formats already configured]}
        "account_issues": account_issues,   # {account: reason string | null}
        "known": known,                     # {account: {ou_number, business_unit, …}}
        # Echoed back so the wizard renders fan-out vs pick-one from one source of
        # truth rather than re-deriving the rule client-side.
        "rows_span_accounts": body.rows_span_accounts,
    }


# ── test a draft recipe ──────────────────────────────────────────────────────────

class BuilderTestRequest(BaseModel):
    storage_key: str
    config_draft: dict          # the draft recipe (source/fields/credit_rule/account_locator/…)


@router.post("/builder/test")
def builder_test(body: BuilderTestRequest, db: Session = Depends(get_db),
                 user: User = Depends(require_permission("config:manage"))):
    """Test a draft recipe against the uploaded file. Returns up to 50 normalized rows."""
    _record, local_path = _local_path_by_key(db, body.storage_key)
    return _test_recipe(local_path, body.config_draft, body.config_draft.get("key", "_DRAFT_"), _record.filename)


@router.post("/builder/infer-date-format")
def builder_infer_date_format(payload: dict,
                              user: User = Depends(require_permission("config:manage"))):
    """Detect the date format from a sample of the mapped Date column's raw
    values (see bank_statement/date_inference.py). Called by the Column Mapping
    step: resolves automatically when the data proves the order, and returns
    'ambiguous' with the competing interpretations only when it genuinely can't
    tell — so the wizard prompts the SPOC. Stateless: takes samples, not a file."""
    from ..bank_statement.date_inference import infer_date_format
    samples = payload.get("samples") or []
    return infer_date_format([str(s) for s in samples])


# ── OU/BU picklist for the wizard's OU step ────────────────────────────────────

@router.get("/builder/available-ous")
def available_ous(db: Session = Depends(get_db),
                  user: User = Depends(require_permission("run:view"))):
    """
    OUs the wizard's OU/Business Unit step can offer, so onboarding an
    account picks from a real, known OU instead of free-typing one that
    might not exist. Two sources, merged:
      1. OUs already onboarded (OrganizationUnit table) — these have a
         known business_unit name + functional_currency already.
      2. OU numbers seen in the currently loaded aging report (the
         authoritative Oracle feed — aging_store's AgingMap.ou_numbers())
         that AREN'T onboarded yet. These show up with business_unit=None
         so the wizard prompts for the name once, on first use.
    """
    known: dict[str, dict] = {}
    for ou in db.query(OrganizationUnit).filter(OrganizationUnit.active.is_(True)).all():
        known[ou.ou_number] = {
            "ou_number": ou.ou_number,
            "business_unit": ou.ou_name,
            "functional_currency": ou.functional_currency,
            "known": True,
        }

    aging_map = aging_store.get_aging_map()
    if aging_map is not None:
        for ou_number in aging_map.ou_numbers:
            if ou_number and ou_number not in known:
                known[ou_number] = {
                    "ou_number": ou_number,
                    "business_unit": None,
                    "functional_currency": None,
                    "known": False,
                }

    return {"ous": sorted(known.values(), key=lambda o: o["ou_number"])}


# ── save a recipe (create account or add a format recipe under an existing one) ──

class AccountAssignment(BaseModel):
    """One account to onboard against the recipe being saved.

    A COLUMN account-locator normally finds several accounts in one file. Each
    needs its own BankAccount + OU (they can differ per account) but shares the
    SAME recipe body, so the wizard sends one of these per discovered account and
    builder_save writes N identical AccountConfigRecipe rows.
    """
    account_number: str
    display_name: str
    ou_number: str
    business_unit: str
    functional_currency: str | None = None
    bank: str | None = None
    currency: str | None = None
    override_account_validation: bool = False


class SaveRecipeRequest(BaseModel):
    account_number: str
    display_name: str
    format: str                       # xlsx | xls | csv | pdf
    recipe: dict                      # account_locator + source + fields + credit_rule + …
    bank: str | None = None
    currency: str | None = None
    # OU + Business Unit are now REQUIRED — every account config must be
    # linked to a real OrganizationUnit, never saved "OU unknown". This is
    # what makes OU/BU a relationship (BankAccount.ou_id -> OrganizationUnit)
    # instead of an optional free-text afterthought.
    ou_number: str
    business_unit: str
    # Ledger/functional currency for this OU — required for Oracle FX Leg 2
    # resolution (rule_engine/fx_service.py's get_functional_currency()).
    # REQUIRED whenever ou_number is genuinely new (see
    # _get_or_create_organization_unit) — there is no fallback/default
    # anymore (it used to silently default to "USD" if left blank, which is
    # exactly the kind of wrong-and-invisible data this field exists to
    # prevent). Not required when ou_number already exists — an existing
    # OU's currency, once set, is never overwritten by a later save.
    functional_currency: str | None = None
    # Explicit SPOC override of the account-number structural gate. False by
    # default: a value that fails account_reject_reason() is hard-blocked at
    # save (it's the join key that corrupts detection across configs when
    # wrong). The wizard only sets this True when the user ticks the
    # "I confirm this is the real account number" checkbox for a rare
    # legitimately-unusual account.
    override_account_validation: bool = False
    # MULTI-ACCOUNT FAN-OUT. When the account locator is a COLUMN, one file
    # legitimately contains several accounts; configuring only one leaves the
    # rest unrecognised (and, before detect_config started refusing such files,
    # got their rows attributed to whichever account won first-fit). The wizard
    # sends every discovered account here — same recipe body, own account
    # number / display name / OU each — and all of them are written in ONE
    # transaction. When omitted, the top-level account_number/display_name/
    # ou_number/business_unit fields are used as a single assignment exactly as
    # before, so existing callers are unaffected.
    accounts: list[AccountAssignment] | None = None
    storage_key: str | None = None    # (unused for keying; kept for symmetry)
    # Best-effort author of this version, read from the login_user_email_stub
    # cookie by the wizard (this module's axios has no dev-user interceptor).
    # Displayed as "added by" in the read-only version list; omitted if unknown.
    created_by: str | None = None


def _get_or_create_organization_unit(db: Session, ou_number: str, business_unit: str,
                                       functional_currency: str | None) -> OrganizationUnit:
    ou_number = ou_number.strip()
    business_unit = business_unit.strip()
    ou = db.query(OrganizationUnit).filter(OrganizationUnit.ou_number == ou_number).first()
    if ou is None:
        # PATCH: this used to silently default to `functional_currency or "USD"`
        # -- meaning an OU created without an explicit currency (e.g. the
        # wizard's field left blank, or a caller forgetting the field) got
        # stamped USD with no warning, no error, nothing in the UI to flag
        # it. Since functional_currency drives real Oracle FX Leg 2
        # resolution (rule_engine/fx_service.py's get_functional_currency())
        # and, once set, is NOT updated on subsequent saves (only ou_name
        # is kept in sync below), a wrong default here is effectively
        # permanent until someone finds and fixes it by hand in the DB.
        # A brand-new OU now REQUIRES an explicit functional_currency from
        # the caller -- no fallback. An OU that already exists doesn't hit
        # this branch at all, so re-saving an existing account never needs
        # to re-supply it.
        if not functional_currency or not functional_currency.strip():
            raise AppError(
                ErrorCode.CONFIG_FIELD_REQUIRED,
                detail=f"Functional Currency -- OU '{ou_number}' is new and has no currency on record yet",
            )
        ou = OrganizationUnit(
            ou_number=ou_number,
            ou_name=business_unit,
            functional_currency=functional_currency.strip().upper(),
        )
        db.add(ou)
        db.flush()
    elif business_unit and ou.ou_name != business_unit:
        # Onboarding this account corrected/updated the BU name for an
        # already-known OU — keep it in sync rather than silently ignoring
        # what the person just typed.
        ou.ou_name = business_unit
    return ou


def _get_or_create_bank_account(db: Session, acct: str, display_name: str, bank: str | None,
                                  currency: str | None,
                                  ou: OrganizationUnit) -> tuple[BankAccount, bool, str | None]:
    """Returns (bank_account, created, previous_ou_number_if_reassigned).

    The third element exists so a fan-out over several accounts can REPORT which
    already-configured accounts had their OU changed by this save. Reassigning an
    existing account's OU is legitimate (it's how a mis-mapped account gets
    corrected) but doing it silently across N accounts at once is not.
    """
    existing = (
        db.query(BankAccount)
        .filter(BankAccount.account_number == acct, BankAccount.bank_name == (bank or "UNKNOWN"))
        .first()
    )
    if existing:
        existing.display_name = display_name
        existing.account_last4 = last4(acct)
        if currency:
            existing.currency = currency
        previous = None
        if existing.ou_id != ou.id:
            prior = existing.organization_unit
            previous = prior.ou_number if prior else None
            existing.ou_id = ou.id
        return existing, False, previous

    account = BankAccount(
        ou_id=ou.id,
        account_number=acct,
        account_last4=last4(acct),
        display_name=display_name,
        bank_name=bank or "UNKNOWN",
        bank_config_key=acct,
        currency=currency,
    )
    db.add(account)
    db.flush()
    return account, True, None


@router.post("/builder/save")
def builder_save(body: SaveRecipeRequest, db: Session = Depends(get_db),
                 user: User = Depends(require_permission("config:manage"))):
    """Save/attach a recipe. If the account exists, a new recipe version is added
    for the format; otherwise a new account is created — always linked to a real
    OrganizationUnit (OU + Business Unit), never saved without one."""
    fmt = _FMT_ALIASES.get(body.format.lower(), body.format.lower())
    if fmt not in ("xlsx", "xls", "csv", "pdf"):
        raise AppError(ErrorCode.CONFIG_FILE_TYPE_UNSUPPORTED, detail=f"'{body.format}' -- use xlsx, xls, csv, or pdf")
    if not body.recipe.get("account_locator"):
        raise AppError(ErrorCode.CONFIG_RECIPE_INVALID, detail="no account locator -- go back to the Account step")

    # One assignment per account to onboard. `accounts` is the multi-account
    # fan-out; without it the top-level fields are a single assignment (the
    # original behaviour).
    if body.accounts:
        assignments = list(body.accounts)
    else:
        assignments = [AccountAssignment(
            account_number=body.account_number,
            display_name=body.display_name,
            ou_number=body.ou_number,
            business_unit=body.business_unit,
            functional_currency=body.functional_currency,
            bank=body.bank,
            currency=body.currency,
            override_account_validation=body.override_account_validation,
        )]

    # Validate EVERY assignment before writing anything, so a bad entry at
    # position 5 of 7 can't leave a half-configured file behind.
    seen_accounts: set[str] = set()
    for a in assignments:
        a.account_number = str(a.account_number).strip()
        if not a.account_number:
            raise AppError(ErrorCode.CONFIG_FIELD_REQUIRED, detail="account number")
        # Structural gate on the account identity — the fix for the config-
        # corruption class of bug. A value that doesn't look like an account
        # (label cell, no digits, too short/long) is refused unless the SPOC
        # explicitly overrode it. Applied per account: with a column locator the
        # discovered set can include a "TOTAL" footer row, and letting one
        # through would register junk as a real account identity.
        if not a.override_account_validation:
            reason = account_reject_reason(a.account_number)
            if reason:
                raise AppError(ErrorCode.CONFIG_RECIPE_INVALID, detail=reason)
        if not (a.display_name or "").strip():
            raise AppError(ErrorCode.CONFIG_FIELD_REQUIRED,
                           detail=f"display name for account {a.account_number}")
        if not (a.ou_number or "").strip():
            raise AppError(ErrorCode.CONFIG_FIELD_REQUIRED,
                           detail=f"Organization Unit for account {a.account_number}")
        if not (a.business_unit or "").strip():
            raise AppError(ErrorCode.CONFIG_FIELD_REQUIRED,
                           detail=f"Business Unit for account {a.account_number}")
        mk = match_key(a.account_number)
        if mk in seen_accounts:
            raise AppError(ErrorCode.CONFIG_RECIPE_INVALID,
                           detail=f"account {a.account_number} is listed twice")
        seen_accounts.add(mk)

    acct = assignments[0].account_number   # back-compat: the primary account

    # Standardize the account/statement currency to an ISO-4217 code (Fusion
    # requires it — see bank_statement/currency.py). Store the ISO on the
    # BankAccount and stamp it into the recipe so the parser has a per-config
    # fallback when a row's own currency value can't be mapped at ingest time.
    if body.currency and body.currency.strip():
        iso = normalize_currency(body.currency)
        if iso is None:
            raise AppError(
                ErrorCode.CONFIG_RECIPE_INVALID,
                detail=f"Currency '{body.currency}' isn't a recognised currency — pick a standard ISO code (e.g. EUR, GBP, USD).",
            )
        body.currency = iso
        body.recipe["currency"] = iso
    # Functional currency (OU ledger) — normalize when it maps cleanly; the OU
    # helper still uppercases/validates presence for a genuinely new OU.
    if body.functional_currency and body.functional_currency.strip():
        body.functional_currency = normalize_currency(body.functional_currency) or body.functional_currency

    # Per-assignment currencies: fall back to the shared statement currency, and
    # reject an unmappable one here rather than letting it reach the DB.
    for a in assignments:
        if a.currency and a.currency.strip():
            iso = normalize_currency(a.currency)
            if iso is None:
                raise AppError(
                    ErrorCode.CONFIG_RECIPE_INVALID,
                    detail=f"Currency '{a.currency}' for account {a.account_number} isn't a recognised currency.",
                )
            a.currency = iso
        else:
            a.currency = body.currency
        if a.functional_currency and a.functional_currency.strip():
            a.functional_currency = normalize_currency(a.functional_currency) or a.functional_currency
        if a.bank is None:
            a.bank = body.bank

    saved: list[dict] = []
    ou_changed: list[dict] = []
    try:
        # ONE transaction for every account — a failure part-way leaves nothing
        # behind, so a multi-account file is never half-configured.
        for a in assignments:
            ou = _get_or_create_organization_unit(db, a.ou_number, a.business_unit, a.functional_currency)
            account, created, previous_ou = _get_or_create_bank_account(
                db, a.account_number, a.display_name, a.bank, a.currency, ou
            )
            if previous_ou is not None:
                ou_changed.append({
                    "account_number": a.account_number,
                    "from_ou_number": previous_ou,
                    "to_ou_number": ou.ou_number,
                })

            existing_versions = (
                db.query(AccountConfigRecipe)
                .filter(AccountConfigRecipe.bank_account_id == account.id,
                        AccountConfigRecipe.format == fmt)
                .all()
            )
            format_created = not existing_versions
            next_version = max((v.version for v in existing_versions), default=0) + 1

            db.add(AccountConfigRecipe(
                bank_account_id=account.id,
                format=fmt,
                version=next_version,
                recipe=body.recipe,          # same recipe body for every account
                created_by=(body.created_by or "").strip() or None,
            ))
            # Flush per account so the next iteration's version query sees this
            # row — matters when the same account somehow appears twice, and
            # keeps the id available for the audit entry.
            db.flush()

            action = "config.create" if (created or format_created) else "config.version_added"
            log_activity(db, user, action=action, entity_type="AccountConfig",
                         entity_id=a.account_number,
                         metadata={"display_name": a.display_name, "format": fmt,
                                   "version": next_version, "ou_number": ou.ou_number,
                                   "business_unit": ou.ou_name,
                                   "batch_size": len(assignments)})

            saved.append({
                "account_number": a.account_number,
                "display_name": a.display_name,
                "created": created,
                "format_created": format_created,
                "version": next_version,
                "ou_number": ou.ou_number,
                "business_unit": ou.ou_name,
            })

        db.commit()
    except AppError:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.exception("builder_save failed for account(s) %s", [a.account_number for a in assignments])
        raise AppError(ErrorCode.CONFIG_SAVE_FAILED, detail=f"{e.__class__.__name__} -- nothing was changed")

    primary = saved[0]
    if len(saved) > 1:
        created_count = sum(1 for s in saved if s["created"])
        message = (
            f"Configured {len(saved)} accounts with the same {fmt} recipe "
            f"({created_count} new, {len(saved) - created_count} updated)."
        )
    elif primary["created"]:
        message = f"Created account {primary['account_number']} with {fmt} recipe (v1)."
    elif primary["format_created"]:
        message = f"Added {fmt} recipe to account {primary['account_number']} (v1)."
    else:
        message = (f"Added version {primary['version']} to {fmt} recipe for account "
                   f"{primary['account_number']}.")

    return {
        "success": True,
        # Back-compat scalars describe the PRIMARY (first) account, so existing
        # callers that ignore `saved` behave exactly as before.
        "account_number": primary["account_number"],
        "format": fmt,
        "created": primary["created"],
        "format_created": primary["format_created"],
        "appended": not primary["created"] and not primary["format_created"],
        "version": primary["version"],
        "ou_number": primary["ou_number"],
        "business_unit": primary["business_unit"],
        "saved": saved,                 # one entry per configured account
        "ou_changed": ou_changed,       # existing accounts whose OU this save moved
        "message": message,
    }


# ── list / fetch accounts (Manage + Clone-from-existing) ─────────────────────────

@router.get("/builder/accounts")
def list_accounts(user: User = Depends(require_permission("run:view"))):
    """List all account configs (light) for the manage dialog and clone picker."""
    out = []
    for acct, entry in load_account_configs().items():
        out.append({
            "account_number": entry.get("account_number", acct),
            "account_last4":  entry.get("account_last4"),
            "display_name":   entry.get("display_name", acct),
            "bank":           entry.get("bank"),
            "currency":       entry.get("currency"),
            # Per-format version metadata (newest first, active flagged) for the
            # read-only version list on the Config tab. Recipe bodies are omitted
            # here — display is metadata-only.
            "formats":        format_summaries(entry),
        })
    return {"accounts": out}


@router.get("/builder/account/{account_number}")
def get_account(account_number: str, user: User = Depends(require_permission("run:view"))):
    """Full account entry (incl. recipes) — used to clone an existing config."""
    entry = load_account_configs().get(account_number)
    if not entry:
        raise AppError(ErrorCode.CONFIG_NOT_FOUND, detail=f"account '{account_number}'")
    return entry


# ── test an existing saved recipe ────────────────────────────────────────────────

class TestExistingRequest(BaseModel):
    filename: str
    account_number: str
    format: str | None = None


@router.post("/test-existing")
def test_existing(body: TestExistingRequest, db: Session = Depends(get_db),
                  user: User = Depends(require_permission("config:manage"))):
    """Run a saved account recipe against the uploaded file (preview + row count)."""
    entry = load_account_configs().get(body.account_number)
    if not entry:
        raise AppError(ErrorCode.CONFIG_NOT_FOUND, detail=f"account '{body.account_number}'")
    record, local_path = _local_path(db, body.filename)
    fmt = body.format or _file_format(body.filename)
    recipe = active_recipe(entry, fmt)   # test always runs the active (latest) version
    if not recipe:
        raise AppError(ErrorCode.CONFIG_NOT_FOUND, detail=f"account '{body.account_number}' has no '{fmt}' recipe")
    return _test_recipe(local_path, recipe, body.account_number, record.filename)


# ── delete account / recipe ──────────────────────────────────────────────────────

@router.delete("/builder/{account_number}")
def delete_account(account_number: str, db: Session = Depends(get_db),
                   user: User = Depends(require_permission("config:manage"))):
    """Delete an entire account config (all its recipes). The OU itself is left
    alone — other accounts may still reference it."""
    account = db.query(BankAccount).filter(BankAccount.account_number == account_number).first()
    if not account:
        raise AppError(ErrorCode.CONFIG_NOT_FOUND, detail=f"account '{account_number}'")
    db.query(AccountConfigRecipe).filter(AccountConfigRecipe.bank_account_id == account.id).delete()
    db.delete(account)
    db.commit()
    reload_account_configs()
    return {"success": True, "deleted": account_number}


@router.delete("/builder/{account_number}/{fmt}")
def delete_recipe(account_number: str, fmt: str, db: Session = Depends(get_db),
                  user: User = Depends(require_permission("config:manage"))):
    """Delete every version of a single format recipe. If it was the account's
    last format, the account itself is removed too."""
    fmt = _FMT_ALIASES.get(fmt.lower(), fmt.lower())
    account = db.query(BankAccount).filter(BankAccount.account_number == account_number).first()
    if not account:
        raise AppError(ErrorCode.CONFIG_NOT_FOUND, detail=f"account '{account_number}'")
    versions = (
        db.query(AccountConfigRecipe)
        .filter(AccountConfigRecipe.bank_account_id == account.id, AccountConfigRecipe.format == fmt)
        .all()
    )
    if not versions:
        raise AppError(ErrorCode.CONFIG_NOT_FOUND,
                        detail=f"account '{account_number}' has no '{fmt}' recipe to delete")
    for v in versions:
        db.delete(v)
    db.flush()
    remaining = (
        db.query(AccountConfigRecipe).filter(AccountConfigRecipe.bank_account_id == account.id).count()
    )
    account_removed = remaining == 0
    if account_removed:
        db.delete(account)
    db.commit()
    reload_account_configs()
    return {"success": True, "account_number": account_number, "format": fmt,
            "account_removed": account_removed}


# ── detect / candidates ──────────────────────────────────────────────────────────

@router.get("/detect/{filename}")
def detect_for_file(filename: str, db: Session = Depends(get_db),
                    user: User = Depends(require_permission("run:view"))):
    """Re-run account-based detection and list matching accounts (picker/reconfigure)."""
    record, local_path = _local_path(db, filename)
    from ..bank_statement.detector import detect_config, list_matching_configs
    det = detect_config(local_path)
    return {
        "filename": filename,
        "success": det.success,
        "config_key": det.config_key,           # matched account number
        "account_number": det.account_number,
        "reason": det.reason,
        "method_detail": det.method_detail,
        "ambiguous": det.reason == "AMBIGUOUS",
        "candidates": list_matching_configs(local_path),
    }