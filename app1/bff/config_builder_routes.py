"""
app.bff.config_builder_routes — /api/config/*  (Account-Based Ingestion)
========================================================================
Config Builder + account-config management for the account-based engine.

  GET    /builder/raw-preview/{filename}   — raw cell grid for the wizard
  POST   /builder/locate-account           — live-extract account(s) via a locator draft
  POST   /builder/test                     — test a *draft recipe* against a file
  POST   /builder/save                     — save a recipe (create account / add format recipe)
  GET    /builder/accounts                 — list accounts (Manage + Clone-from-existing)
  GET    /builder/account/{account_number} — full account entry (for cloning)
  DELETE /builder/{account_number}          — delete an account (all its recipes)
  DELETE /builder/{account_number}/{format} — delete a single format recipe
  POST   /test-existing                    — test a *saved* recipe against a file
  GET    /detect/{filename}                — re-detect + list matching accounts

Config store is JSON only: account_configs.json (+ account_ou_map.json).
Register under /api/config alongside config_routes.
"""
from __future__ import annotations

import os
import json
import dataclasses
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..db.models import SourceFile
from ..deps import get_db
from ..storage.client import get_storage_client
from ..bank_statement.account_locator import extract_accounts, normalize_account, last4, match_key
from ..bank_statement.configs.account_loader import (
    load_account_configs, load_account_ou_map, reload_account_configs,
)

router = APIRouter()

_CFG_DIR = Path(__file__).parent.parent / "bank_statement" / "configs"
_ACCOUNT_CONFIGS_PATH = _CFG_DIR / "account_configs.json"
_OU_MAP_PATH = _CFG_DIR / "account_ou_map.json"
_STATEMENT_BUCKET = "bank-statements"
_FMT_ALIASES = {"xlsm": "xlsx", "txt": "csv"}


# ── shared helpers ────────────────────────────────────────────────────────────

def _local_path(db: Session, filename: str) -> tuple[SourceFile, str]:
    record = db.query(SourceFile).filter(
        SourceFile.kind == "bank_statement", SourceFile.filename == filename
    ).first()
    if not record:
        raise HTTPException(404, f"File '{filename}' not found")
    storage = get_storage_client()
    return record, storage.local_path_for_read(_STATEMENT_BUCKET, record.storage_key)


def _local_path_by_key(db: Session, storage_key: str) -> tuple[SourceFile, str]:
    record = db.query(SourceFile).filter(
        SourceFile.kind == "bank_statement", SourceFile.storage_key == storage_key
    ).first()
    if not record:
        raise HTTPException(404, "File not found")
    storage = get_storage_client()
    return record, storage.local_path_for_read(_STATEMENT_BUCKET, record.storage_key)


def _file_format(filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower().lstrip(".")
    return _FMT_ALIASES.get(ext, ext)


def _read_raw(path: Path) -> dict:
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _test_recipe(local_path: str, recipe: dict, key: str, filename: str) -> dict:
    """Parse a file with a recipe and return a row-count + sample preview."""
    from ..bank_statement.detector import DetectionResult
    from ..bank_statement.parser import parse_credit_rows, ColumnValidationError

    detection = DetectionResult(success=True, config_key=key, config=recipe,
                                step_used=1, method_detail=f"Test '{key}'")
    try:
        rows = parse_credit_rows(local_path, detection, filename)

        def _ser(r):
            d = dataclasses.asdict(r)
            if d.get("statement_date") is not None:
                d["statement_date"] = str(d["statement_date"])
            return d

        return {"success": True, "row_count": len(rows), "rows": [_ser(r) for r in rows[:50]]}
    except (ColumnValidationError, Exception) as e:
        return {"success": False, "error": str(e), "row_count": 0, "rows": []}


# ── raw preview ─────────────────────────────────────────────────────────────────

@router.get("/builder/raw-preview/{filename}")
def builder_raw_preview(filename: str, db: Session = Depends(get_db)):
    """Return raw cell data for all sheets — used by the wizard preview/header/locate steps."""
    record, local_path = _local_path(db, filename)
    ext = os.path.splitext(filename)[1].lower().lstrip(".")

    MAX_ROWS, MAX_COLS = 40, 20
    sheets = []
    if ext in ("xlsx", "xlsm"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(local_path, read_only=True, data_only=True)
            for sh in wb.sheetnames:
                ws = wb[sh]
                rows = [[str(v).strip() if v is not None else "" for v in row]
                        for row in ws.iter_rows(min_row=1, max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True)]
                sheets.append({"name": sh, "rows": rows})
            wb.close()
        except Exception as e:
            raise HTTPException(500, f"Could not read xlsx: {e}")
    elif ext == "xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(local_path)
            for sh in wb.sheet_names():
                ws = wb.sheet_by_name(sh)
                rows = []
                for r in range(min(MAX_ROWS, ws.nrows)):
                    row = [str(ws.cell_value(r, c)).strip() if ws.cell_value(r, c) not in (None, "") else ""
                           for c in range(min(MAX_COLS, ws.ncols))]
                    row += [""] * (MAX_COLS - len(row))
                    rows.append(row)
                sheets.append({"name": sh, "rows": rows})
        except Exception as e:
            raise HTTPException(500, f"Could not read xls: {e}")
    elif ext in ("csv", "txt"):
        try:
            import csv
            with open(local_path, encoding="utf-8", errors="replace") as f:
                rows = [[str(v).strip() for v in row[:MAX_COLS]]
                        for i, row in enumerate(csv.reader(f)) if i < MAX_ROWS]
            sheets.append({"name": "Sheet1", "rows": rows})
        except Exception as e:
            raise HTTPException(500, f"Could not read csv: {e}")
    else:
        raise HTTPException(400, f"Unsupported extension: .{ext}")

    return {"filename": filename, "storage_key": record.storage_key, "extension": ext, "sheets": sheets}


# ── locate account (live preview for the wizard's account step) ──────────────────

class LocateAccountRequest(BaseModel):
    storage_key: str
    locator: dict
    source: dict | None = None


@router.post("/builder/locate-account")
def builder_locate_account(body: LocateAccountRequest, db: Session = Depends(get_db)):
    """Run a locator draft against the file and return the account(s) it finds,
    flagging any that are already registered."""
    _record, local_path = _local_path_by_key(db, body.storage_key)
    found = sorted(extract_accounts(local_path, body.locator, body.source or {}))

    configs = load_account_configs()
    existing = {}
    for acct in found:
        mk = match_key(acct)
        for k, entry in configs.items():
            if match_key(entry.get("account_number", k)) == mk:
                existing[acct] = sorted((entry.get("recipes") or {}).keys())
    return {
        "accounts": found,
        "count": len(found),
        "last4s": sorted({last4(a) for a in found}),
        "existing": existing,   # {account: [formats already configured]}
    }


# ── test a draft recipe ──────────────────────────────────────────────────────────

class BuilderTestRequest(BaseModel):
    storage_key: str
    config_draft: dict          # the draft recipe (source/fields/credit_rule/account_locator/…)


@router.post("/builder/test")
def builder_test(body: BuilderTestRequest, db: Session = Depends(get_db)):
    """Test a draft recipe against the uploaded file. Returns up to 50 normalized rows."""
    _record, local_path = _local_path_by_key(db, body.storage_key)
    return _test_recipe(local_path, body.config_draft, body.config_draft.get("key", "_DRAFT_"), _record.filename)


# ── save a recipe (create account or add a format recipe under an existing one) ──

class SaveRecipeRequest(BaseModel):
    account_number: str
    display_name: str
    format: str                       # xlsx | xls | csv | pdf
    recipe: dict                      # account_locator + source + fields + credit_rule + …
    bank: str | None = None
    currency: str | None = None
    ou_number: str | None = None
    business_unit: str | None = None
    storage_key: str | None = None    # (unused for keying; kept for symmetry)


@router.post("/builder/save")
def builder_save(body: SaveRecipeRequest):
    """Save/attach a recipe. If the account exists, the recipe is added (or replaced)
    under recipes[format]; otherwise a new account entry is created. Validates the
    whole registry after writing and rolls back on failure."""
    acct = str(body.account_number).strip()
    if not acct:
        raise HTTPException(400, "Account number is required")
    if not body.display_name.strip():
        raise HTTPException(400, "Display name is required")
    fmt = _FMT_ALIASES.get(body.format.lower(), body.format.lower())
    if fmt not in ("xlsx", "xls", "csv", "pdf"):
        raise HTTPException(400, f"Unsupported format '{body.format}'")
    if not body.recipe.get("account_locator"):
        raise HTTPException(400, "Recipe must include an account_locator")

    raw = _read_raw(_ACCOUNT_CONFIGS_PATH)
    before = json.dumps(raw)

    entry = raw.get(acct) if not acct.startswith("_") else None
    overwritten = False
    created = False
    if isinstance(entry, dict):
        recipes = entry.setdefault("recipes", {})
        overwritten = fmt in recipes
        recipes[fmt] = body.recipe
        entry["display_name"] = body.display_name
        entry["account_last4"] = last4(acct)
        if body.bank:     entry["bank"] = body.bank
        if body.currency: entry["currency"] = body.currency
        entry["account_number"] = acct
    else:
        created = True
        raw[acct] = {
            "account_number": acct,
            "account_last4":  last4(acct),
            "bank":           body.bank or "",
            "currency":       body.currency or "",
            "display_name":   body.display_name,
            "recipes":        {fmt: body.recipe},
        }

    # write + validate, roll back on failure
    with open(_ACCOUNT_CONFIGS_PATH, "w", encoding="utf-8") as f:
        json.dump(raw, f, indent=2)
    try:
        reload_account_configs()
    except Exception as e:
        with open(_ACCOUNT_CONFIGS_PATH, "w", encoding="utf-8") as f:
            f.write(before)
        reload_account_configs()
        raise HTTPException(400, f"Config invalid, not saved: {e}")

    # optional OU mapping
    if body.ou_number:
        ou = _read_raw(_OU_MAP_PATH)
        ou[acct] = {"ou_number": body.ou_number, "business_unit": body.business_unit or ""}
        with open(_OU_MAP_PATH, "w", encoding="utf-8") as f:
            json.dump(ou, f, indent=2)
        reload_account_configs()

    return {
        "success": True,
        "account_number": acct,
        "format": fmt,
        "created": created,
        "overwritten": overwritten,
        "message": (f"Added {fmt} recipe to existing account {acct}."
                    if not created else f"Created account {acct} with {fmt} recipe."),
    }


# ── list / fetch accounts (Manage + Clone-from-existing) ─────────────────────────

@router.get("/builder/accounts")
def list_accounts():
    """List all account configs (light) for the manage dialog and clone picker."""
    out = []
    for acct, entry in load_account_configs().items():
        out.append({
            "account_number": entry.get("account_number", acct),
            "account_last4":  entry.get("account_last4"),
            "display_name":   entry.get("display_name", acct),
            "bank":           entry.get("bank"),
            "currency":       entry.get("currency"),
            "formats":        sorted((entry.get("recipes") or {}).keys()),
        })
    return {"accounts": out}


@router.get("/builder/account/{account_number}")
def get_account(account_number: str):
    """Full account entry (incl. recipes) — used to clone an existing config."""
    entry = load_account_configs().get(account_number)
    if not entry:
        raise HTTPException(404, f"Account '{account_number}' not found")
    return entry


# ── test an existing saved recipe ────────────────────────────────────────────────

class TestExistingRequest(BaseModel):
    filename: str
    account_number: str
    format: str | None = None


@router.post("/test-existing")
def test_existing(body: TestExistingRequest, db: Session = Depends(get_db)):
    """Run a saved account recipe against the uploaded file (preview + row count)."""
    entry = load_account_configs().get(body.account_number)
    if not entry:
        raise HTTPException(404, f"Account '{body.account_number}' not found")
    record, local_path = _local_path(db, body.filename)
    fmt = body.format or _file_format(body.filename)
    recipe = (entry.get("recipes") or {}).get(fmt)
    if not recipe:
        raise HTTPException(404, f"Account '{body.account_number}' has no '{fmt}' recipe")
    return _test_recipe(local_path, recipe, body.account_number, record.filename)


# ── delete account / recipe ──────────────────────────────────────────────────────

@router.delete("/builder/{account_number}")
def delete_account(account_number: str):
    """Delete an entire account config (all its recipes) + its OU mapping."""
    raw = _read_raw(_ACCOUNT_CONFIGS_PATH)
    if account_number not in raw or account_number.startswith("_"):
        raise HTTPException(404, f"Account '{account_number}' not found")
    del raw[account_number]
    with open(_ACCOUNT_CONFIGS_PATH, "w", encoding="utf-8") as f:
        json.dump(raw, f, indent=2)

    ou = _read_raw(_OU_MAP_PATH)
    if ou.pop(account_number, None) is not None:
        with open(_OU_MAP_PATH, "w", encoding="utf-8") as f:
            json.dump(ou, f, indent=2)

    reload_account_configs()
    return {"success": True, "deleted": account_number}


@router.delete("/builder/{account_number}/{fmt}")
def delete_recipe(account_number: str, fmt: str):
    """Delete a single format recipe. If it was the last recipe, the account is removed."""
    fmt = _FMT_ALIASES.get(fmt.lower(), fmt.lower())
    raw = _read_raw(_ACCOUNT_CONFIGS_PATH)
    entry = raw.get(account_number)
    if not isinstance(entry, dict) or fmt not in (entry.get("recipes") or {}):
        raise HTTPException(404, f"Account '{account_number}' has no '{fmt}' recipe")
    del entry["recipes"][fmt]
    account_removed = False
    if not entry["recipes"]:
        del raw[account_number]
        account_removed = True
    with open(_ACCOUNT_CONFIGS_PATH, "w", encoding="utf-8") as f:
        json.dump(raw, f, indent=2)
    reload_account_configs()
    return {"success": True, "account_number": account_number, "format": fmt,
            "account_removed": account_removed}


# ── detect / candidates ──────────────────────────────────────────────────────────

@router.get("/detect/{filename}")
def detect_for_file(filename: str, db: Session = Depends(get_db)):
    """Re-run account-based detection and list matching accounts (picker/reconfigure)."""
    record, local_path = _local_path(db, filename)
    from ..bank_statement.detector import detect_config, list_matching_configs
    det = detect_config(local_path)
    return {
        "filename": filename,
        "success": det.success,
        "config_key": det.config_key,           # matched account number
        "account_number": det.account_number,
        "reason": det.reason,
        "method_detail": det.method_detail,
        "ambiguous": det.reason == "AMBIGUOUS",
        "candidates": list_matching_configs(local_path),
    }