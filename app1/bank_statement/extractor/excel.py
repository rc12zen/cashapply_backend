"""
app.bank_statement.extractor.excel
=====================================
ExcelExtractor handles .xlsx (openpyxl) and .xls (xlrd), with support for
single-row and multi-row merged headers (split Dr/Cr pattern).
"""
from __future__ import annotations

import pandas as pd


def _detect_excel_engine(filepath: str) -> str:
    """Determine xlrd vs openpyxl from file magic bytes, falling back to extension."""
    try:
        with open(filepath, "rb") as f:
            magic = f.read(8)
        if magic[:4] == b"\xd0\xcf\x11\xe0":  # OLE2 compound document (xls)
            return "xlrd"
        if magic[:2] == b"PK":  # ZIP archive (xlsx/xlsm)
            return "openpyxl"
    except Exception:
        pass
    return "xlrd" if filepath.lower().endswith(".xls") else "openpyxl"


def _resolve_sheet(filepath: str, sheet_cfg: dict, engine: str) -> str | int:
    by = sheet_cfg.get("by", "first")
    value = sheet_cfg.get("value")

    if by == "index":
        return int(value)
    if by == "first":
        return 0
    if by == "name":
        return value
    if by == "name_contains":
        if engine == "xlrd":
            import xlrd
            wb = xlrd.open_workbook(filepath)
            names = wb.sheet_names()
        else:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            names = list(wb.sheetnames)
            wb.close()
        v_lower = str(value).lower()
        for name in names:
            if v_lower in name.lower():
                return name
        raise ValueError(f"No sheet containing '{value}' found in {filepath}")
    raise ValueError(f"Unknown sheet selector: '{by}'")


def _read_merged_header(filepath: str, sheet, header_cfg: dict, engine: str) -> pd.DataFrame:
    """Read a DataFrame where the header spans two (or more) rows.

    Reads header rows directly (bypassing pandas MultiIndex forward-fill) so that
    empty header cells remain empty rather than inheriting the previous column's label.
    """
    main_row = header_cfg["row"]
    merge_rows_cfg = header_cfg.get("merge_rows", [])
    sub_row_nums = [mr["row"] for mr in merge_rows_cfg]

    # --- Step 1: read raw header values without pandas forward-fill ---
    if engine == "openpyxl":
        from openpyxl import load_workbook
        _wb = load_workbook(filepath, read_only=True, data_only=True)
        _ws = _wb[sheet] if isinstance(sheet, str) else list(_wb.worksheets)[int(sheet)]

        def _raw_row(row_0based: int) -> list:
            rows = list(_ws.iter_rows(
                min_row=row_0based + 1, max_row=row_0based + 1, values_only=True,
            ))
            if not rows:
                return []
            return [str(v).strip() if v is not None else "" for v in rows[0]]

        top_vals = _raw_row(main_row)
        sub_vals = [_raw_row(r) for r in sub_row_nums]
        _wb.close()
    else:
        import xlrd
        _wb_xls = xlrd.open_workbook(filepath)
        _ws_xls = (
            _wb_xls.sheet_by_name(sheet)
            if isinstance(sheet, str)
            else _wb_xls.sheet_by_index(int(sheet))
        )

        def _raw_row(row_0based: int) -> list:  # type: ignore[misc]
            if row_0based >= _ws_xls.nrows:
                return []
            return [
                str(_ws_xls.cell_value(row_0based, c)).strip()
                for c in range(_ws_xls.ncols)
            ]

        top_vals = _raw_row(main_row)
        sub_vals = [_raw_row(r) for r in sub_row_nums]

    # --- Step 2: build merge rules and flat column names ---
    merge_rules: dict[str, str] = {}
    for mr_cfg in merge_rows_cfg:
        for rule in mr_cfg.get("rules", []):
            merge_rules[rule["sub_value"].lower()] = rule["rename_parent_to"]

    first_sub = sub_vals[0] if sub_vals else []
    flat_cols: list[str] = []
    for i, top in enumerate(top_vals):
        sub = first_sub[i] if i < len(first_sub) else ""
        renamed = merge_rules.get(sub.lower())
        flat_cols.append(renamed if renamed else (top or sub or f"_col_{i}"))

    # --- Step 3: read data rows only (skip all header rows) ---
    last_header_row = max([main_row] + sub_row_nums)
    df = pd.read_excel(
        filepath,
        sheet_name=sheet,
        header=None,
        skiprows=list(range(last_header_row + 1)),
        engine=engine,
    )

    # Align columns (trim extra trailing columns, apply flat names)
    n = min(len(flat_cols), len(df.columns))
    df = df.iloc[:, :n]
    df.columns = flat_cols[:n]
    return df


class ExcelExtractor:
    @staticmethod
    def extract(filepath: str, source_cfg: dict) -> pd.DataFrame:
        engine = _detect_excel_engine(filepath)
        sheet_cfg = source_cfg.get("sheet", {"by": "first"})
        sheet = _resolve_sheet(filepath, sheet_cfg, engine)
        header_cfg = source_cfg.get("header", {"row": 0})

        if "merge_rows" in header_cfg:
            df = _read_merged_header(filepath, sheet, header_cfg, engine)
        else:
            df = pd.read_excel(
                filepath, sheet_name=sheet, header=header_cfg["row"], engine=engine
            )
            df.columns = [str(c).strip() for c in df.columns]

        # Drop blank-header columns (pandas auto-names them "Unnamed: N")
        df = df.loc[:, ~df.columns.str.startswith("Unnamed:")]
        return df
